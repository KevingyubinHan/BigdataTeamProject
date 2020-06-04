# card keyword extraction.py
import os
import pandas as pd
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows

# read card_summary (csv->df)
csv_path = r'D:\Classification Rule\Regular Expression\04. 데이터그룹화 추가작업\raw data'
csv_name = '_Card_summary_info_cat_200309.csv'
card_summary_df = pd.read_csv(csv_path + os.sep + csv_name)
print(card_summary_df.shape)

# create excel workbook
wb = xl.Workbook()
workbook_path = r'D:\Classification Rule\Regular Expression\04. 데이터그룹화 추가작업\200508_가전제품'
workbook_name = 'keyword_extraction_200508.xlsx'
wb.save(filename=workbook_path + os.sep + workbook_name)

# load excel workbook
wb = xl.load_workbook(workbook_path + os.sep + workbook_name)


# dataframe to excel
def df_to_excel(df, ws):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


def change_df_to_excel(wb, excel_name, sheets_list, df_list):
    for i in range(len(sheets_list)):
        ws = wb[sheets_list[i]]
        df_to_excel(df_list[i], ws)
        wb.save(filename=excel_name)


# shop_name(str->list->excel sheet naming)
def name_worksheet(wb, names_list):
    for i in range(len(names_list)):
        if i == 0:
            ws = wb.active
            ws.title = names_list[i]
        else:
            ws = wb.create_sheet(title=names_list[i])
    wb.save(filename=workbook_path + os.sep + workbook_name)


shop_names_str = '''에스원
세스코
이투스
메가스터디
대성마이맥
스카이에듀
박문각
에듀윌'''
shop_names_list = shop_names_str.split('\n')
name_worksheet(wb, shop_names_list)
shop_df_list = []

# extract keyword in card_summary
extraction_df = card_summary_df.loc[
                (card_summary_df['desc'].str.contains('^APL') |
                 card_summary_df['desc'].str.contains('^APPLE')), :]

print(extraction_df.shape)
print(extraction_df['desc'])

shop_df_list.append(extraction_df)
shop_df_list.insert(1, extraction_df)

# save excel after extraction
change_df_to_excel(wb, workbook_path + os.sep + workbook_name, shop_names_list, shop_df_list)

# dataframe to excel(in convenience)
wb = xl.load_workbook(workbook_path + os.sep + workbook_name)
ws = wb.create_sheet()
ws.title = '엠베스트'

df_to_excel(extraction_df, ws)
wb.save(filename=workbook_path + os.sep + workbook_name)

# data sort & drop duplicate
data_directory = r'D:\Classification Rule\Regular Expression\03. 정규식 match 결과'
data_name = r'RegExpr_hit_total_200402_카드,편의서비스,대출.xlsx'
data_path = data_directory + os.sep + data_name
data_df = pd.read_excel(io=data_path)

data_df.sort_values(by=['desc', 'rule_index'], inplace=True, ascending=True)
data_df.drop_duplicates(subset='desc', keep='first', inplace=True)
